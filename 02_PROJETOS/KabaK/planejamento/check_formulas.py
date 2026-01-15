
import openpyxl

file_path = r'C:\Users\Gassen\OneDrive\Meu_Segundo_Cerebro\02_PROJETOS\KabaK\planejamento\PLANILHA_KABAK_SANSOM.xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Sheet1']
    
    # Check Lucro Mensal (Row 11 in CSV -> Row 12 in Excel likely)
    # Custo Estabilidade is Row 4 in CSV -> Row 5 in Excel
    
    cell_profit = ws.cell(row=12, column=3) # Mai/26
    print(f"Cell {cell_profit.coordinate} value: {cell_profit.value}")
    
    cell_cost = ws.cell(row=5, column=3) # Mai/26
    print(f"Cell {cell_cost.coordinate} value: {cell_cost.value}")

except Exception as e:
    print(f"Error: {e}")
