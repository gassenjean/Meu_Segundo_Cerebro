
import openpyxl
from openpyxl.styles import Font

file_path = r'C:\Users\Gassen\OneDrive\Meu_Segundo_Cerebro\02_PROJETOS\KabaK\planejamento\PLANILHA_KABAK_SANSOM.xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Sheet1']
    
    # 1. Update Header with Societal Info
    ws['A1'] = "PROJETO KABAK - 50% Sansom / 50% Jean+Gassen+Kris"
    ws['A1'].font = Font(bold=True, color="0000FF")

    # 2. Iterate Columns B (2) to N (14) -> 13 months? CSV had 13 cols (0-12). Excel has A-M?
    # CSV cols: 1 to 12 (Apr/26 to Mar/27) -> 12 months.
    # Let's verify range. CSV had header empty, then Abr/26...
    # Row 1 in Excel: A=Empty, B=Abr/26, C=Mai/26... M=Mar/27?
    # Let's count columns.
    
    # We will iterate from Col 2 (B) to max column.
    max_col = ws.max_column
    
    accumulated_cash = 0
    
    for col in range(2, max_col + 1):
        # Rows (1-based):
        # 3: Receita
        # 4: Custo Estabilidade (To Zero)
        # 5: Custos Var
        # 6: Mkt Trafego
        # 7: Mkt Titanium
        # 8: Logistica
        # 9: Operacao
        # 10: Impostos
        # 11: Lucro Mensal (Calc)
        # 12: Caixa Acum (Calc)
        # 2: Investimento (Calc)
        
        # Zero out Stability
        ws.cell(row=5, column=col).value = 0
        
        # Get values
        rev = ws.cell(row=4, column=col).value or 0
        
        # Expenses
        c_stab = 0 # Forced
        c_var = ws.cell(row=6, column=col).value or 0
        c_mkt1 = ws.cell(row=7, column=col).value or 0
        c_mkt2 = ws.cell(row=8, column=col).value or 0
        c_log = ws.cell(row=9, column=col).value or 0
        c_op = ws.cell(row=10, column=col).value or 0
        c_tax = ws.cell(row=11, column=col).value or 0
        
        total_expenses = c_var + c_mkt1 + c_mkt2 + c_log + c_op + c_tax
        
        # Profit
        profit = rev - total_expenses
        ws.cell(row=12, column=col).value = profit
        
        # Accumulated Cash (Previous + Profit + Investment)
        # We need Investment first.
        # Logic: If profit < 0, Investment covers it?
        # Let's see original logic.
        # Original: Inv 580k, Profit -580k. Cash 0.
        # New: Profit = -80k. Inv should be 80k?
        # If I strictly follow "Draft" cleanup, keeping logic simple.
        
        inv = 0
        if profit < 0:
            inv = abs(profit)
        
        ws.cell(row=3, column=col).value = inv
        
        accumulated_cash += (profit + inv)
        ws.cell(row=13, column=col).value = accumulated_cash

    wb.save(file_path)
    print("Excel updated successfully.")

except Exception as e:
    print(f"Error: {e}")
