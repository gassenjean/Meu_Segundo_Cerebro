
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- CONFIGURACAO DO CRONOGRAMA ---
# Feedback: "estabilidade seria por 6 meses... valores que vamos precisar"
# Fonte: PROPOSTA_FINAL_KABAK_SANSOM.md
# Custeio Fabricao KABAK: R$ 500.000/ms x 6 meses (Estabilidade)

months = [
    "Abr/26", "Mai/26", "Jun/26", "Jul/26", "Ago/26", "Set/26",
    "Out/26", "Nov/26", "Dez/26", "Jan/27", "Fev/27", "Mar/27"
]

data = {
    "Investimento": [0.0] * 12,
    "Receita": [0.0] * 12,
    "Custo Estabilidade": [0.0] * 12, # Novo Item: Custeio Fabrica
    "Custos Var. Venda": [0.0] * 12, # Apenas variveis da venda (impostos, taxa, etc)
    "Mkt Trfego": [0.0] * 12,
    "Mkt Titanium": [0.0] * 12,
    "Logstica": [0.0] * 12,
    "Operao Fixo": [0.0] * 12,
    "Impostos": [0.0] * 12,
    "Lucro Mensal": [0.0] * 12,
    "Caixa Acum.": [0.0] * 12
}

# --- PREMISSAS ---
preco_venda = 129.00
imposto_pct = 0.025
gateway_pct = 0.03
logistica_pct = 0.12 # Varivel sobre venda
custo_kit_reposicao = 48.00 # Custo para repor estoque APS o buffer (se necessrio)

# --- FLUXO MENSAL ---

# MESES 1-6 (ABR-SET): FASE ESTABILIDADE + ESCALA
# Fabrica recebe R$ 500k/ms FIXO para cobrir custos e produzir estoque.
# Investimento Sansom cobre isso.

for i in range(6): # Abr a Set
    data["Custo Estabilidade"][i] = 500000.0 # Custo fixo da fbrica garantido
    # Isso  financiado por investimento
    data["Investimento"][i] += 500000.0 

# MARKETING E OPERACIONAL TITANIUM
# Abr: Setup 60k
# Mai em diante: 60k + Trafego
mkt_titanium_fee = 60000.0
trafego_inicial = 40000.0
trafego_escala = [40000, 40000, 50000, 70000, 90000, 100000] # Mai-Out

# ABRIL (Ms 0)
idx = 0
data["Mkt Titanium"][idx] = mkt_titanium_fee
data["Operao Fixo"][idx] = 20000.0 
# Investimento cobre Mkt + Op + Estabilidade
custo_extra_abr = data["Mkt Titanium"][idx] + data["Operao Fixo"][idx]
data["Investimento"][idx] += custo_extra_abr 

# MAIO a SETEMBRO (Ms 1-5)
vendas_proj = [1000, 3000, 8000, 12000, 15000] # Mai, Jun, Jul, Ago, Set
# Nota: O custo do produto vendido (CPV) est sendo coberto pelo "Custo Estabilidade" de 500k/ms?
# SIM. O documento diz "Fabricao: R$ 500k/ms (Sansom paga) ... Produo 80k peas".
# Ento NO abatemos custo de kit varivel nestes meses, pois j est pago no fixo de estabilidade.
# Abatemos apenas Impostos, Logstica e Gateway.

for m_rel in range(5):
    idx = m_rel + 1 # Comea em Maio (idx 1)
    qtd = vendas_proj[m_rel]
    rec = qtd * preco_venda
    data["Receita"][idx] = rec
    
    # Variveis de venda
    var_venda = rec * (imposto_pct + gateway_pct + logistica_pct)
    data["Custos Var. Venda"][idx] = var_venda
    
    # Marketing
    data["Mkt Titanium"][idx] = mkt_titanium_fee
    data["Mkt Trfego"][idx] = trafego_escala[m_rel]
    
    # Operacional/Logistica Fixa
    data["Logstica"][idx] = 10000.0 + (m_rel * 2000)
    data["Operao Fixo"][idx] = 40000.0 + (m_rel * 2000)
    data["Impostos"][idx] = rec * imposto_pct
    
    # Investimento Sansom cobre Marketing tambm nos primeiros meses?
    # Proposta diz: "Investimento Sansom (3 meses): Fabricao 1.5M + Mkt 300k".
    # Ento sim, investimento cobre o deficit de caixa nesses meses iniciais.
    
    pass

# MESES 7-12 (OUT-MAR): AUTO-SUSTENTAVEL
# Custo Estabilidade ACABA. Fabrica passa a cobrar por pea ou mantm custo?
# Proposta: "Ms 7+: Marca paga prpria produo... Sansom no investe mais".
# Ento voltamos ao modelo de Custo Variabilizado (R$ 48/kit) ou Custo Fabrica absorvido.
# Para segurana, vamos considerar Custo Varivel de reposio.

vendas_estavel = 18000
for i in range(6, 12): # Out a Mar
    data["Receita"][i] = vendas_estavel * preco_venda
    
    data["Custo Estabilidade"][i] = 0.0 # Fim da estabilidade garantida
    
    # Agora pagamos o produto por unidade (reposio)
    cpv = vendas_estavel * custo_kit_reposicao
    var_venda = data["Receita"][i] * (imposto_pct + gateway_pct + logistica_pct)
    data["Custos Var. Venda"][i] = cpv + var_venda
    
    data["Mkt Titanium"][i] = mkt_titanium_fee
    data["Mkt Trfego"][i] = 100000.0
    
    data["Logstica"][i] = 30000.0
    data["Operao Fixo"][i] = 60000.0
    data["Impostos"][i] = data["Receita"][i] * imposto_pct

# --- CALCULO CAIXA ---
caixa_acumulado = 0.0
for i in range(12):
    despesas_totais = (data["Custo Estabilidade"][i] + data["Custos Var. Venda"][i] + 
                       data["Mkt Trfego"][i] + data["Mkt Titanium"][i] + 
                       data["Logstica"][i] + data["Operao Fixo"][i] + 
                       data["Impostos"][i])
    
    data["Lucro Mensal"][i] = data["Receita"][i] - despesas_totais
    
    # Fluxo Caixa = Lucro + Investimento
    saldo = data["Lucro Mensal"][i] + data["Investimento"][i]
    caixa_acumulado += saldo
    
    # Se caixa negativo (mesmo com aporte da estabilidade), cobrimos a diferena se for fase 1
    if caixa_acumulado < 0 and i < 6:
        deficit = abs(caixa_acumulado)
        data["Investimento"][i] += deficit
        caixa_acumulado += deficit
    
    data["Caixa Acum."][i] = caixa_acumulado

# --- EXPORT ---
df = pd.DataFrame(data, index=months).T
output_file = r"C:\Users\gasse\OneDrive\Meu_Segundo_Cerebro\02_PROJETOS\KabaK\planejamento\PLANILHA_KABAK_SANSOM.xlsx"
df.to_excel(output_file)

# Estilizao
wb = openpyxl.load_workbook(output_file)
ws = wb.active

header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")

for row in ws.iter_rows(min_row=2):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = u'R$ #,##0.00;[Red]-R$ #,##0.00'

wb.save(output_file)
print("Sucesso: Planilha v5 (Com Estabilidade R$ 500k/mes) gerada.")
