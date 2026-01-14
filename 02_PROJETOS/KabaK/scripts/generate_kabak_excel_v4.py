
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- CONFIGURACAO DO CRONOGRAMA ---
# Feedback: "começar a planilha no mes que ira iniciAR O INVESTIMENTO"
# Inicio: Abril/2026
# Periodo: 12 meses (Abril/26 a Maro/27)

months = [
    "Abr/26", "Mai/26", "Jun/26", "Jul/26", "Ago/26", "Set/26",
    "Out/26", "Nov/26", "Dez/26", "Jan/27", "Fev/27", "Mar/27"
]

# Inicializar dados zerados
data = {
    "Investimento": [0.0] * 12,
    "Receita": [0.0] * 12,
    "Custos Var.": [0.0] * 12,
    "Mkt Trfego": [0.0] * 12,
    "Mkt Titanium": [0.0] * 12,
    "Logstica": [0.0] * 12,
    "Operao Fixo": [0.0] * 12,
    "Impostos": [0.0] * 12,
    "Lucro Mensal": [0.0] * 12,
    "Caixa Acum.": [0.0] * 12
}

# --- PREMISSAS ---
custo_kit = 48.00
preco_venda = 129.00
imposto_pct = 0.025
gateway_pct = 0.03
logistica_pct = 0.12
logistica_fixa = 10000.0
op_fixo_base = 40000.0

# --- FLUXO MENSAL ---

# MES 1 - ABRIL: PRE-OPERACIONAL / INVESTIMENTO
# ndice 0
idx_abr = 0
data["Mkt Titanium"][idx_abr] = 60000.0  # Setup
data["Operao Fixo"][idx_abr] = 20000.0   # Equipe parcial
estoque_inicial_custo = 400000.0         # Compra tecido
data["Custos Var."][idx_abr] = estoque_inicial_custo 

# Investimento Abril:
custo_abr = data["Mkt Titanium"][idx_abr] + data["Operao Fixo"][idx_abr] + data["Custos Var."][idx_abr]
data["Investimento"][idx_abr] = custo_abr + 100000.0 # +100k capital giro
# Total Investimento Abril ~580k

# MES 2 - MAIO: INICIO VENDAS (1000 kits)
idx_mai = 1
vendas_mai = 1000
rec_mai = vendas_mai * preco_venda
data["Receita"][idx_mai] = rec_mai
var_venda_mai = rec_mai * (imposto_pct + gateway_pct + logistica_pct)
cpv_mai = vendas_mai * custo_kit 
data["Custos Var."][idx_mai] = cpv_mai + var_venda_mai
data["Mkt Trfego"][idx_mai] = 40000.0
data["Mkt Titanium"][idx_mai] = 60000.0
data["Logstica"][idx_mai] = logistica_fixa
data["Operao Fixo"][idx_mai] = op_fixo_base
data["Impostos"][idx_mai] = rec_mai * imposto_pct

# MES 3 - JUNHO: 3000 kits
idx_jun = 2
vendas_jun = 3000
rec_jun = vendas_jun * preco_venda
data["Receita"][idx_jun] = rec_jun
var_venda_jun = rec_jun * (imposto_pct + gateway_pct + logistica_pct)
cpv_jun = vendas_jun * custo_kit
data["Custos Var."][idx_jun] = cpv_jun + var_venda_jun # +100k reforo estoque removido para simplificar modelo visual
data["Mkt Trfego"][idx_jun] = 50000.0
data["Mkt Titanium"][idx_jun] = 60000.0
data["Logstica"][idx_jun] = logistica_fixa
data["Operao Fixo"][idx_jun] = op_fixo_base
data["Impostos"][idx_jun] = rec_jun * imposto_pct

# MES 4 - JULHO: 8000 kits
idx_jul = 3
vendas_jul = 8000
rec_jul = vendas_jul * preco_venda
data["Receita"][idx_jul] = rec_jul
var_venda_jul = rec_jul * (imposto_pct + gateway_pct + logistica_pct)
cpv_jul = vendas_jul * custo_kit
data["Custos Var."][idx_jul] = cpv_jul + var_venda_jul
data["Mkt Trfego"][idx_jul] = 70000.0
data["Mkt Titanium"][idx_jul] = 60000.0
data["Logstica"][idx_jul] = logistica_fixa + 5000.0
data["Operao Fixo"][idx_jul] = op_fixo_base + 5000.0
data["Impostos"][idx_jul] = rec_jul * imposto_pct

# MES 5 - AGOSTO: 12000 kits
idx_ago = 4
vendas_ago = 12000
rec_ago = vendas_ago * preco_venda
data["Receita"][idx_ago] = rec_ago
var_venda_ago = rec_ago * (imposto_pct + gateway_pct + logistica_pct)
cpv_ago = vendas_ago * custo_kit
data["Custos Var."][idx_ago] = cpv_ago + var_venda_ago
data["Mkt Trfego"][idx_ago] = 90000.0
data["Mkt Titanium"][idx_ago] = 60000.0
data["Logstica"][idx_ago] = logistica_fixa + 10000.0
data["Operao Fixo"][idx_ago] = op_fixo_base + 10000.0
data["Impostos"][idx_ago] = rec_ago * imposto_pct

# MES 6 - SETEMBRO: 15000 kits
idx_set = 5
vendas_set = 15000
rec_set = vendas_set * preco_venda
data["Receita"][idx_set] = rec_set
var_venda_set = rec_set * (imposto_pct + gateway_pct + logistica_pct)
cpv_set = vendas_set * custo_kit
data["Custos Var."][idx_set] = cpv_set + var_venda_set
data["Mkt Trfego"][idx_set] = 100000.0
data["Mkt Titanium"][idx_set] = 60000.0
data["Logstica"][idx_set] = logistica_fixa + 15000.0
data["Operao Fixo"][idx_set] = op_fixo_base + 15000.0
data["Impostos"][idx_set] = rec_set * imposto_pct

# MESES 7 a 12 (OUT-MAR): 18000 kits estabilizado
for i in range(6, 12):
    vendas_plat = 18000
    rec_plat = vendas_plat * preco_venda
    data["Receita"][i] = rec_plat
    var_venda_plat = rec_plat * (imposto_pct + gateway_pct + logistica_pct)
    cpv_plat = vendas_plat * custo_kit
    data["Custos Var."][i] = cpv_plat + var_venda_plat
    data["Mkt Trfego"][i] = 100000.0
    data["Mkt Titanium"][i] = 60000.0
    data["Logstica"][i] = logistica_fixa + 20000.0
    data["Operao Fixo"][i] = op_fixo_base + 20000.0
    data["Impostos"][i] = rec_plat * imposto_pct

# --- CALCULO LUCRO E CAIXA ---
caixa_acumulado = 0.0
for i in range(12):
    despesas_totais = (data["Custos Var."][i] + data["Mkt Trfego"][i] + 
                       data["Mkt Titanium"][i] + data["Logstica"][i] + 
                       data["Operao Fixo"][i] + data["Impostos"][i])
    
    lucro = data["Receita"][i] - despesas_totais
    data["Lucro Mensal"][i] = lucro
    
    # Investimento entra no caixa
    saldo_mes = lucro + data["Investimento"][i]
    caixa_acumulado += saldo_mes
    
    # Se caixa ficar negativo, aporta diferena
    if caixa_acumulado < 0:
        aporte_extra = abs(caixa_acumulado) + 50000.0 
        data["Investimento"][i] += aporte_extra
        caixa_acumulado += aporte_extra
        
    data["Caixa Acum."][i] = caixa_acumulado

# --- EXPORT ---
df = pd.DataFrame(data, index=months).T
output_file = r"C:\Users\gasse\OneDrive\Meu_Segundo_Cerebro\02_PROJETOS\KabaK\planejamento\PLANILHA_KABAK_SANSOM.xlsx"
df.to_excel(output_file)

# Estilizao visual
wb = openpyxl.load_workbook(output_file)
ws = wb.active

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
currency_format = u'R$ #,##0.00;[Red]-R$ #,##0.00'

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for row in ws.iter_rows(min_row=2):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = currency_format

for column_cells in ws.columns:
    try:
        max_len = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_len + 2
    except:
        pass

wb.save(output_file)
print("Sucesso: Planilha v4 (12 meses a partir de Abril) gerada.")
