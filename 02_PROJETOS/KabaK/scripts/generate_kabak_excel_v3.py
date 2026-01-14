
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURAO DO CRONOGRAMA ---
# Feedback do usurio: "falou que nao ia investir logo em janeiro"
# Transcrio: Equipe volta em maro, foco total em Abril.
# Titanium: Abril. Vendas: Maio.

months = [
    "Jan/26", "Fev/26", "Mar/26", "Abr/26", "Mai/26", "Jun/26",
    "Jul/26", "Ago/26", "Set/26", "Out/26", "Nov/26", "Dez/26"
]

# Inicializar dados zerados
data = {
    "Investimento": [0.0] * 12,
    "Receita": [0.0] * 12,
    "Custos Var.": [0.0] * 12, # Fabricao + Impostos variveis da venda
    "Mkt Trfego": [0.0] * 12,
    "Mkt Titanium": [0.0] * 12,
    "Logstica": [0.0] * 12,
    "Operao Fixo": [0.0] * 12,
    "Impostos": [0.0] * 12, # Simples/Outros
    "Lucro Mensal": [0.0] * 12,
    "Caixa Acum.": [0.0] * 12
}

# --- PREMISSAS ---
custo_kit = 48.00  # R$ 48,00 (30 tecido + 15 fabricao + 3 embalagem)
preco_venda = 129.00
imposto_pct = 0.025 # 2.5%
gateway_pct = 0.03 # 3%
logistica_pct = 0.12 # 12%
logistica_fixa = 10000.0 # Fixo mensal RS 10kCD
mkt_titanium = 60000.0
op_fixo_base = 40000.0 # Operacional base (equipe, etc)
investimento_inicial_pecas = 1000000.0 # Investimento "Bala" para estoque inicial (Buffer)

# --- FLUXO MENSAL ---

# JAN / FEV / MAR: Zero ou custos baixssimos (Ignorado conforme pedido "no investir em jan")

# ABRIL: PRE-OPERACIONAL "REAL"
# Investimento forte para preparar Maio.
# Custos: Titanium Setup, Estoque Inicial (Tecido/Produo), Marketing Setup.
idx_abr = 3
data["Mkt Titanium"][idx_abr] = 60000.0  # Setup
data["Operao Fixo"][idx_abr] = 20000.0   # Meia equipe/preparao
stoque_inicial_custo = 400000.0          # Compra tecido/produo antecipada
data["Custos Var."][idx_abr] = stoque_inicial_custo 

# Investimento Abril: Cobrir custos + Caixa Mnimo
custo_abr = data["Mkt Titanium"][idx_abr] + data["Operao Fixo"][idx_abr] + data["Custos Var."][idx_abr]
data["Investimento"][idx_abr] = custo_abr + 100000.0 # +100k capital giro inicial
# Total Investimento Abril ~580k

# MAIO: INICIO VENDAS (1000 kits)
idx_mai = 4
vendas_mai = 1000
rec_mai = vendas_mai * preco_venda
data["Receita"][idx_mai] = rec_mai
# Custos variveis de VENDA (impostos, taxa, logistica var)
var_venda_mai = rec_mai * (imposto_pct + gateway_pct + logistica_pct)
# Custo do produto vendido (CPV) - J pago no estoque ou novo?
# Vamos considerar reposio contnua para simplificar fluxo caixa
cpv_mai = vendas_mai * custo_kit 
data["Custos Var."][idx_mai] = cpv_mai + var_venda_mai
data["Mkt Trfego"][idx_mai] = 40000.0
data["Mkt Titanium"][idx_mai] = 60000.0
data["Logstica"][idx_mai] = logistica_fixa # Fixo logistica
data["Operao Fixo"][idx_mai] = op_fixo_base
data["Impostos"][idx_mai] = rec_mai * imposto_pct

# JUNHO: 3000 kits
idx_jun = 5
vendas_jun = 3000
rec_jun = vendas_jun * preco_venda
data["Receita"][idx_jun] = rec_jun
var_venda_jun = rec_jun * (imposto_pct + gateway_pct + logistica_pct)
cpv_jun = vendas_jun * custo_kit
data["Custos Var."][idx_jun] = cpv_jun + var_venda_jun + 100000.0 # +100k reforo estoque
data["Mkt Trfego"][idx_jun] = 50000.0
data["Mkt Titanium"][idx_jun] = 60000.0
data["Logstica"][idx_jun] = logistica_fixa
data["Operao Fixo"][idx_jun] = op_fixo_base
data["Impostos"][idx_jun] = rec_jun * imposto_pct

# JULHO: 8000 kits (Break-even?)
idx_jul = 6
vendas_jul = 8000
rec_jul = vendas_jul * preco_venda
data["Receita"][idx_jul] = rec_jul
var_venda_jul = rec_jul * (imposto_pct + gateway_pct + logistica_pct)
cpv_jul = vendas_jul * custo_kit
data["Custos Var."][idx_jul] = cpv_jul + var_venda_jul
data["Mkt Trfego"][idx_jul] = 70000.0
data["Mkt Titanium"][idx_jul] = 60000.0
data["Logstica"][idx_jul] = logistica_fixa + 5000.0 # Escala logistica
data["Operao Fixo"][idx_jul] = op_fixo_base + 5000.0
data["Impostos"][idx_jul] = rec_jul * imposto_pct

# AGOSTO: 12000 kits
idx_ago = 7
vendas_ago = 12000
rec_ago = vendas_ago * preco_venda
data["Receita"][idx_ago] = rec_ago
var_venda_ago = rec_ago * (imposto_pct + gateway_pct + logistica_pct)
cpv_ago = vendas_ago * custo_kit
data["Custos Var."][idx_ago] = cpv_ago + var_venda_ago
data["Mkt Trfego"][idx_ago] = 90000.0
data["Mkt Titanium"][idx_ago] = 60000.0
data["Logstica"][idx_ago] = logistica_fixa + 10000.0
data["Operao Fixo"][idx_ago] = op_fixo_base + 10000.0
data["Impostos"][idx_ago] = rec_ago * imposto_pct

# SETEMBRO: 15000 kits
idx_set = 8
vendas_set = 15000
rec_set = vendas_set * preco_venda
data["Receita"][idx_set] = rec_set
var_venda_set = rec_set * (imposto_pct + gateway_pct + logistica_pct)
cpv_set = vendas_set * custo_kit
data["Custos Var."][idx_set] = cpv_set + var_venda_set
data["Mkt Trfego"][idx_set] = 100000.0
data["Mkt Titanium"][idx_set] = 60000.0
data["Logstica"][idx_set] = logistica_fixa + 15000.0
data["Operao Fixo"][idx_set] = op_fixo_base + 15000.0
data["Impostos"][idx_set] = rec_set * imposto_pct

# OUT - DEZ: 18000 kits (Plat)
for i in range(9, 12):
    vendas_plat = 18000
    rec_plat = vendas_plat * preco_venda
    data["Receita"][i] = rec_plat
    var_venda_plat = rec_plat * (imposto_pct + gateway_pct + logistica_pct)
    cpv_plat = vendas_plat * custo_kit
    data["Custos Var."][i] = cpv_plat + var_venda_plat
    data["Mkt Trfego"][i] = 100000.0
    data["Mkt Titanium"][i] = 60000.0
    data["Logstica"][i] = logistica_fixa + 20000.0
    data["Operao Fixo"][i] = op_fixo_base + 20000.0
    data["Impostos"][i] = rec_plat * imposto_pct

# --- CALCULO LUCRO E CAIXA ---
caixa_acumulado = 0.0
for i in range(12):
    # Despesas Totais
    despesas_totais = (data["Custos Var."][i] + data["Mkt Trfego"][i] + 
                       data["Mkt Titanium"][i] + data["Logstica"][i] + 
                       data["Operao Fixo"][i] + data["Impostos"][i])
    
    # Lucro (Operacional) = Receita - Despesas
    # Nota: No subtramos o investimento aqui, pois Investimento  Capital Social (Entrada de Caixa)
    # Mas para o "Lucro" DRE, Investimento no conta.
    # Porm para o Caixa, conta.
    
    lucro = data["Receita"][i] - despesas_totais
    data["Lucro Mensal"][i] = lucro
    
    # Check Caixa
    # Saldo Ms = Lucro + Investimento (Investimento cobre o buraco)
    saldo_mes = lucro + data["Investimento"][i]
    
    caixa_acumulado += saldo_mes
    
    # Se caixa ficar negativo em Maio/Jun, precisamos de mais investimento
    if caixa_acumulado < 0 and i >= 3: # A partir de Abril
        deficit = abs(caixa_acumulado)
        # Adiciona investimento para zerar e deixar um buffer
        aporte_extra = deficit + 50000.0 
        data["Investimento"][i] += aporte_extra
        caixa_acumulado += aporte_extra
        
    data["Caixa Acum."][i] = caixa_acumulado

# --- PANDAS & EXCEL ---
df = pd.DataFrame(data, index=months).T # Transpor para ficar Meses nas Colunas

output_file = r"C:\Users\gasse\OneDrive\Meu_Segundo_Cerebro\02_PROJETOS\KabaK\planejamento\PLANILHA_KABAK_SANSOM.xlsx"
df.to_excel(output_file)

# Beautify Excel
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
currency_format = u'R$ #,##0.00;[Red]-R$ #,##0.00'

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for row in ws.iter_rows(min_row=2):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = currency_format

# Adjust widths
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

wb.save(output_file)
print("Sucesso: Planilha v3 (Inicio Abril) gerada.")
print(f"DEBUG: Investimento Jan = {data['Investimento'][0]}")
print(f"DEBUG: Investimento Abr = {data['Investimento'][3]}")
