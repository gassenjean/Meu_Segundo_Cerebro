
import openpyxl

file_path = r'C:\Users\Gassen\OneDrive\Meu_Segundo_Cerebro\02_PROJETOS\KabaK\planejamento\PLANILHA_KABAK_SANSOM.xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Sheet1']
    
    # Check Lucro Mensal (Row 11, Col 3 for Mai/26)
    cell = ws.cell(row=11, column=3)
    print(f"Cell {cell.coordinate} value: {cell.value}, data_type: {cell.data_type}")
    
    # Check Investimento (Row 2, Col 3)
    cell_inv = ws.cell(row=2, column=3)
    print(f"Cell {cell_inv.coordinate} value: {cell_inv.value}, data_type: {cell_inv.data_type}")

except Exception as e:
    print(f"Error: {e}")
